"""
Service for generating Excel exports.
"""
import pandas as pd
from datetime import datetime, timedelta
from typing import Dict, Any, List
from sqlalchemy.orm import Session
from database.models import Document, ExtractedField, Match, Mismatch, ClientProfile, Export
from google.cloud import storage
from pydantic_settings import BaseSettings
import os
from dotenv import load_dotenv
import io
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()


class ExportSettings(BaseSettings):
    """Export configuration."""
    gcs_bucket_name: str = os.getenv("GCS_BUCKET_NAME", "")

    class Config:
        env_file = ".env"
        extra = "ignore"  # Ignore extra fields from .env


class ExportService:
    """Service for generating Excel exports."""

    def __init__(self):
        """Initialize export service."""
        self.settings = ExportSettings()
        self.storage_client = storage.Client()

    def generate_excel_report(
        self,
        db: Session,
        doc_id: int
    ) -> Dict[str, Any]:
        """
        Generate Excel report for a document.
        
        Args:
            db: Database session
            doc_id: Document ID
            
        Returns:
            Dictionary with export information
        """
        # Get document
        document = db.query(Document).filter(Document.id == doc_id).first()
        if not document:
            raise ValueError(f"Document {doc_id} not found")
        
        # Get extracted fields
        extracted_fields = db.query(ExtractedField).filter(
            ExtractedField.doc_id == doc_id
        ).all()
        
        # Convert to dictionary
        fields_dict = {}
        for field in extracted_fields:
            fields_dict[field.field_name] = {
                'raw_value': field.raw_value,
                'normalized_value': field.normalized_value,
                'confidence': field.confidence_score
            }
        
        # Get match
        match = db.query(Match).filter(Match.doc_id == doc_id).first()
        matched_client_id = None
        match_score = 0.0
        if match:
            matched_client_id = match.client_id
            match_score = match.match_score
        
        # Get client profile if matched
        expected_dob = None
        expected_doa = None
        if matched_client_id:
            client = db.query(ClientProfile).filter(
                ClientProfile.id == matched_client_id
            ).first()
            if client:
                expected_dob = client.dob.strftime('%Y-%m-%d') if client.dob else None
                expected_doa = client.doa.strftime('%Y-%m-%d') if client.doa else None
        
        # Get mismatches
        mismatches = db.query(Mismatch).filter(Mismatch.doc_id == doc_id).all()
        mismatch_dict = {m.field: {'expected': m.expected_value, 'observed': m.observed_value} for m in mismatches}
        
        # Get client name if matched
        client_name = None
        if matched_client_id:
            client = db.query(ClientProfile).filter(ClientProfile.id == matched_client_id).first()
            if client:
                client_name = client.name
        
        # Prepare extracted field data
        extracted_name = fields_dict.get('patient_name', {}).get('normalized_value', '') or fields_dict.get('patient_name', {}).get('raw_value', '')
        extracted_dob = fields_dict.get('dob', {}).get('normalized_value', '') or fields_dict.get('dob', {}).get('raw_value', '')
        extracted_doa = fields_dict.get('doa', {}).get('normalized_value', '') or fields_dict.get('doa', {}).get('raw_value', '')
        extracted_referral = fields_dict.get('referral', {}).get('normalized_value', '') or fields_dict.get('referral', {}).get('raw_value', '')
        
        # Handle service_dates (might be a list)
        service_dates_field = fields_dict.get('service_dates', {})
        service_dates = service_dates_field.get('normalized_value', '') or service_dates_field.get('raw_value', '')
        if isinstance(service_dates, list):
            service_dates = '; '.join(str(d) for d in service_dates if d)
        
        # Get confidence scores
        name_confidence = fields_dict.get('patient_name', {}).get('confidence', 0)
        dob_confidence = fields_dict.get('dob', {}).get('confidence', 0)
        doa_confidence = fields_dict.get('doa', {}).get('confidence', 0)
        referral_confidence = fields_dict.get('referral', {}).get('confidence', 0)
        
        # Determine match status for each field
        dob_match_status = 'Matched'
        if 'dob' in mismatch_dict:
            dob_match_status = 'Mismatch'
        elif not expected_dob:
            dob_match_status = 'Not in Dataset'
        elif not extracted_dob:
            dob_match_status = 'Not Extracted'
        
        doa_match_status = 'Matched'
        if 'doa' in mismatch_dict:
            doa_match_status = 'Mismatch'
        elif not expected_doa:
            doa_match_status = 'Not in Dataset'
        elif not extracted_doa:
            doa_match_status = 'Not Extracted'
        
        name_match_status = 'Matched' if match and match.decision in ['match', 'ambiguous'] else ('No Match' if match else 'Not Checked')
        
        # Create main results DataFrame
        results_data = {
            'Document ID': [doc_id],
            'File Name': [document.filename],
            'Processed At': [document.updated_at.strftime('%Y-%m-%d %H:%M:%S') if document.updated_at else ''],
            'Match Status': [match.decision if match else 'No Match'],
            'Match Score (%)': [f"{match_score:.1f}" if match_score else "N/A"],
            'Matched Client ID': [matched_client_id if matched_client_id else 'N/A'],
            'Matched Client Name': [client_name if client_name else 'N/A'],
        }
        
        # Create field-by-field DataFrame
        fields_data = {
            'Field Name': [
                'Patient Name',
                'Date of Birth',
                'Date of Accident',
                'Referral Number'
            ],
            'Extracted Value': [
                extracted_name,
                extracted_dob,
                extracted_doa,
                extracted_referral
            ],
            'Expected Value (from Dataset)': [
                client_name if client_name else 'N/A',
                expected_dob if expected_dob else 'Not in Dataset',
                expected_doa if expected_doa else 'Not in Dataset',
                'N/A (Not checked)'
            ],
            'Match Status': [
                name_match_status,
                dob_match_status,
                doa_match_status,
                'N/A'
            ],
            'Confidence (%)': [
                f"{(name_confidence * 100):.1f}" if name_confidence else "N/A",
                f"{(dob_confidence * 100):.1f}" if dob_confidence else "N/A",
                f"{(doa_confidence * 100):.1f}" if doa_confidence else "N/A",
                f"{(referral_confidence * 100):.1f}" if referral_confidence else "N/A"
            ]
        }
        
        # Create mismatches DataFrame if any
        mismatches_data = None
        if mismatches:
            mismatches_data = {
                'Field': [m.field.upper() for m in mismatches],
                'Expected Value': [m.expected_value for m in mismatches],
                'Observed Value': [m.observed_value for m in mismatches],
                'Mismatch Type': ['Date Mismatch' for _ in mismatches]
            }
        
        # Create service dates DataFrame if available
        service_dates_data = None
        if service_dates:
            service_dates_data = {
                'Service Dates': [service_dates]
            }
        
        # Generate Excel file in memory
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # Write Results sheet
            df_results = pd.DataFrame(results_data)
            df_results.to_excel(writer, index=False, sheet_name='Summary')
            
            # Write Fields sheet
            df_fields = pd.DataFrame(fields_data)
            df_fields.to_excel(writer, index=False, sheet_name='Field Matching')
            
            # Write Mismatches sheet if any
            if mismatches_data:
                df_mismatches = pd.DataFrame(mismatches_data)
                df_mismatches.to_excel(writer, index=False, sheet_name='Mismatches')
            
            # Write Service Dates sheet if available
            if service_dates_data:
                df_service = pd.DataFrame(service_dates_data)
                df_service.to_excel(writer, index=False, sheet_name='Service Dates')
        
        # Apply formatting
        excel_buffer.seek(0)
        workbook = load_workbook(excel_buffer)
        
        # Format Summary sheet
        self._format_summary_sheet(workbook['Summary'])
        
        # Format Field Matching sheet
        self._format_fields_sheet(workbook['Field Matching'])
        
        # Format Mismatches sheet if exists
        if 'Mismatches' in workbook.sheetnames:
            self._format_mismatches_sheet(workbook['Mismatches'])
        
        # Format Service Dates sheet if exists
        if 'Service Dates' in workbook.sheetnames:
            self._format_service_dates_sheet(workbook['Service Dates'])
        
        # Save formatted workbook
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        excel_content = excel_buffer.read()
        
        # Try to upload to GCS, fallback to direct download if it fails
        filename = f"exports/report_{doc_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        gcs_uri = None
        signed_url = None
        
        try:
            if self.settings.gcs_bucket_name:
                gcs_uri = self._upload_to_gcs(excel_content, filename)
                signed_url = self._generate_signed_url(filename, expiration_minutes=60)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"⚠️ GCS upload failed, will use direct download: {str(e)}")
            # Continue without GCS - we'll return the file directly
        
        # Save export record (even if GCS failed)
        export = Export(
            doc_id=doc_id,
            gcs_uri=gcs_uri or "N/A",
            signed_url=signed_url or "N/A",
            expires_at=datetime.now() + timedelta(hours=1) if signed_url else None
        )
        db.add(export)
        db.commit()
        
        # If GCS failed, return file content directly
        if not signed_url:
            # Store file content in a way that can be returned
            # We'll use base64 encoding for the response
            import base64
            file_base64 = base64.b64encode(excel_content).decode('utf-8')
            
            return {
                'export_id': export.id,
                'gcs_uri': None,
                'signed_url': None,
                'file_content': file_base64,
                'filename': f"report_{doc_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                'expires_at': None,
                'direct_download': True
            }
        
        return {
            'export_id': export.id,
            'gcs_uri': gcs_uri,
            'signed_url': signed_url,
            'expires_at': export.expires_at.isoformat(),
            'direct_download': False
        }

    def _upload_to_gcs(self, content: bytes, filename: str) -> str:
        """Upload file to GCS."""
        if not self.settings.gcs_bucket_name:
            raise ValueError("GCS bucket name not configured")
        bucket = self.storage_client.bucket(self.settings.gcs_bucket_name)
        blob = bucket.blob(filename)
        blob.upload_from_string(content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        return f"gs://{self.settings.gcs_bucket_name}/{filename}"

    def _generate_signed_url(self, filename: str, expiration_minutes: int = 60) -> str:
        """Generate signed URL for GCS object."""
        if not self.settings.gcs_bucket_name:
            raise ValueError("GCS bucket name not configured")
        bucket = self.storage_client.bucket(self.settings.gcs_bucket_name)
        blob = bucket.blob(filename)
        
        url = blob.generate_signed_url(
            expiration=timedelta(minutes=expiration_minutes),
            method='GET'
        )
        return url
    
    def _format_summary_sheet(self, sheet):
        """Format the Summary sheet."""
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Format data rows
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center')
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_fields_sheet(self, sheet):
        """Format the Field Matching sheet."""
        # Header style
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Format data rows with conditional coloring
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
            match_status_cell = sheet[f'D{row_idx}']  # Match Status column
            match_status = match_status_cell.value
            
            # Color code based on match status
            if match_status == 'Matched':
                row_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif match_status == 'Mismatch':
                row_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif match_status == 'Not in Dataset':
                row_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                row_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            
            for cell in row:
                cell.fill = row_fill
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_mismatches_sheet(self, sheet):
        """Format the Mismatches sheet."""
        # Header style
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Format data rows (red background for mismatches)
        mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.fill = mismatch_fill
                cell.border = border
                cell.alignment = Alignment(vertical='center')
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_service_dates_sheet(self, sheet):
        """Format the Service Dates sheet."""
        # Header style
        header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Format data rows
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

